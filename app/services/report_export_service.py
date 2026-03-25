# app/services/report_export_service.py

import io
import re
import asyncio
import logging
from pathlib import Path
from typing import Dict, Any, List, Optional, Tuple

from sqlalchemy.orm import Session
from openpyxl import load_workbook
from openpyxl.cell.cell import MergedCell
from openpyxl.drawing.image import Image as XLImage
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.styles import Alignment, Font
from app.models.complaint import Complaint
from app.models.report import Report
from app.models.report_step import ReportStep
from app.models.file import File as FileModel
from app.models.step_file import StepFile
from app.services.file_storage import storage  # GitHub singleton

logger = logging.getLogger(__name__)

# ── Template ──────────────────────────────────────────────────────────────────
TEMPLATE_PATH = (
    Path(__file__).parent.parent / "templates" / "Problem_solving_template_2026.xlsx"
)

# ── MIME types we can embed as Excel images ───────────────────────────────────
_IMAGE_MIMES = {
    "image/jpeg",
    "image/jpg",
    "image/png",
    "image/gif",
    "image/bmp",
    "image/tiff",
}
_STEP_ORDER = ["D1", "D2", "D3", "D4", "D5", "D6", "D7", "D8"]
# =============================================================================
# GENERIC CELL HELPERS
# =============================================================================


def _s(value, default: str = "") -> str:
    if value is None:
        return default
    if isinstance(value, bool):
        return "✓" if value else ""
    return str(value).strip()


def _step_data(steps: Dict[str, ReportStep], code: str) -> Dict[str, Any]:
    s = steps.get(code)
    return (s.data or {}) if s else {}


def _safe_filename(name: str) -> str:
    name = name.replace("\u2013", "-").replace("\u2014", "-")
    name = name.replace("\u2018", "'").replace("\u2019", "'")
    name = name.encode("latin-1", errors="ignore").decode("latin-1")
    name = re.sub(r'[\\/*?:"<>|]', "", name)
    return name.strip()


def _resolve_cell(ws: Worksheet, coord: str):
    """
    Return the writable top-left cell for a given coordinate.
    If coord falls inside a merged range, return the master (top-left) cell.
    If coord is already a normal cell, return it directly.
    """
    cell = ws[coord]
    if not isinstance(cell, MergedCell):
        return cell
    for merge_range in ws.merged_cells.ranges:
        if coord in merge_range:
            master = ws.cell(merge_range.min_row, merge_range.min_col)
            if not isinstance(master, MergedCell):
                return master
    return None  # slave with no findable master — skip


def _w(ws: Worksheet, coord: str, value) -> None:
    """Write a value — automatically redirects slave merged cells to their master."""
    cell = _resolve_cell(ws, coord)
    if cell is not None:
        cell.value = value


def _wb(ws: Worksheet, coord: str, flag: bool) -> None:
    """Write a boolean — automatically redirects slave merged cells to their master."""
    cell = _resolve_cell(ws, coord)
    if cell is not None:
        cell.value = bool(flag)


def _wrap(ws: Worksheet, coord: str) -> None:
    cell = _resolve_cell(ws, coord)
    if cell is None:
        return
    ex = cell.alignment or Alignment()
    cell.alignment = Alignment(
        wrap_text=True,
        horizontal=ex.horizontal,
        vertical=ex.vertical,
        indent=ex.indent or 0,
        text_rotation=ex.text_rotation or 0,
    )


def _auto_height(ws: Worksheet, row: int, text: str, col_width: float = 27) -> None:
    if not text:
        return
    lines = text.split("\n")
    total = sum(max(1, int(len(ln) / max(col_width, 1)) + 1) for ln in lines)
    ws.row_dimensions[row].height = min(max(15.0, total * 15.0), 200.0)


def _hyperlink(ws: Worksheet, coord: str, url: str, label: str) -> None:
    """Write a clickable hyperlink into a cell."""
    cell = _resolve_cell(ws, coord)
    if cell is None:
        return
    cell.value = label
    cell.hyperlink = url
    cell.font = Font(color="0563C1", underline="single")


# =============================================================================
# GITHUB FILE HELPERS
# =============================================================================


def _fetch_bytes_sync(stored_name: str) -> Optional[bytes]:
    """Sync wrapper around the async GitHub fetch."""
    try:
        loop = asyncio.new_event_loop()
        try:
            return loop.run_until_complete(storage.fetch_content(stored_name))
        finally:
            loop.close()
    except Exception as exc:
        logger.warning("GitHub fetch failed for '%s': %s", stored_name, exc)
        return None


def _get_step_files(
    db: Session,
    step: ReportStep,
    images_only: bool = False,
) -> List[FileModel]:
    """All files attached to a step, optionally filtered to images only."""
    q = (
        db.query(FileModel)
        .join(StepFile, StepFile.file_id == FileModel.id)
        .filter(StepFile.report_step_id == step.id)
        .order_by(StepFile.attachment_order)
    )
    if images_only:
        q = q.filter(FileModel.mime_type.in_(_IMAGE_MIMES))
    return q.all()


def _categorise_files(
    db: Session,
    step: Optional[ReportStep],
) -> Tuple[Optional[bytes], List[Tuple[str, str]]]:
    """
    For a given step return:
      image_bytes  — bytes of the first embeddable image (or None)
      links        — list of (label, url) for every non-image file
                     PLUS any image files after the first one
    """
    if step is None:
        return None, []

    data = step.data or {}
    evidence = data.get("evidence_documents", "") or ""

    all_files = _get_step_files(db, step, images_only=False)
    file_by_name = {f.original_name: f for f in all_files}

    ordered: List[FileModel] = []
    seen_ids = set()

    # Evidence-named files first
    for fname in [n.strip() for n in evidence.split(",") if n.strip()]:
        f = file_by_name.get(fname)
        if f and f.id not in seen_ids:
            ordered.append(f)
            seen_ids.add(f.id)

    # Remaining attached files
    for f in all_files:
        if f.id not in seen_ids:
            ordered.append(f)
            seen_ids.add(f.id)

    image_bytes: Optional[bytes] = None
    links: List[Tuple[str, str]] = []

    for f in ordered:
        url = storage.url_for(f.stored_path)

        if f.mime_type in _IMAGE_MIMES and image_bytes is None:
            raw = _fetch_bytes_sync(f.stored_path)
            if raw:
                image_bytes = raw
                continue  # embedded — skip link for this one

        links.append((f.original_name, url))

    return image_bytes, links


def _insert_photo(
    ws: Worksheet,
    img_bytes: bytes,
    anchor: str = "G15",
    width: int = 380,
    height: int = 180,
) -> None:
    try:
        img = XLImage(io.BytesIO(img_bytes))
        img.width = width
        img.height = height
        img.anchor = anchor
        ws.add_image(img)
    except Exception as exc:
        logger.warning("Could not embed photo at %s: %s", anchor, exc)


def _write_links(
    ws: Worksheet,
    start_coord: str,
    links: List[Tuple[str, str]],
) -> None:
    """Write one hyperlink per row starting at start_coord."""
    if not links:
        return
    col_letter = re.match(r"([A-Z]+)", start_coord).group(1)
    start_row = int(re.match(r"[A-Z]+(\d+)", start_coord).group(1))
    for i, (label, url) in enumerate(links):
        _hyperlink(ws, f"{col_letter}{start_row + i}", url, f"📎 {label}")


# =============================================================================
# D1 — Team
# =============================================================================


def _member_role_bucket(member: Dict[str, Any]) -> str:
    fn = _s(member.get("function", "")).lower()
    dept = _s(member.get("department", "")).lower()
    blob = f"{fn} {dept}".strip()

    def has(*kw):
        return any(k in blob for k in kw)

    if has(
        "production",
        "operator",
        "line leader",
        "supervisor",
        "line ",
        "shopfloor",
        "shop floor",
    ):
        return "production"
    if has("maintenance", "technician", "mechanic", "electrical", "mechanical"):
        return "maintenance"
    if has(
        "engineering", "process", "manufacturing", "industrial", "methods", "method"
    ):
        return "engineering"
    if has(
        "logistics",
        "supply",
        "warehouse",
        "shipping",
        "receiving",
        "supplier",
        "sqa",
        "procurement",
    ):
        return "logistics"
    if has("team leader", "leader", "project", "pm", "manager", "management"):
        return "leader"
    if has("quality", "qc", "qa", "quality control", "quality assurance"):
        return "quality"
    return "other"


def _fill_d1(ws: Worksheet, data: Dict) -> None:
    ROLE_ROWS = {
        "production": 6,
        "maintenance": 7,
        "engineering": 8,
        "logistics": 9,
        "leader": 10,
        "quality": 11,
        "other": 12,
    }
    buckets: Dict[str, List[str]] = {k: [] for k in ROLE_ROWS}

    for member in data.get("team_members", []) or []:
        line = (
            f"{_s(member.get('name'))}  —  "
            f"{_s(member.get('function'))} / {_s(member.get('department'))}"
        ).strip()
        if not line or line == "— /":
            continue
        buckets[_member_role_bucket(member)].append(line)

    for role, row in ROLE_ROWS.items():
        text = "\n".join(buckets[role])
        _w(ws, f"B{row}", text)
        _wb(ws, f"J{row}", bool(buckets[role]))
        if text:
            _wrap(ws, f"B{row}")
            _auto_height(ws, row, text, col_width=27)


# =============================================================================
# D2 — Problem description + photo + Is/Is-Not
# =============================================================================


def _fill_d2(
    ws: Worksheet,
    data: Dict,
    db: Session,
    d2_step: Optional[ReportStep],
) -> None:
    five_w: Dict = data.get("five_w_2h", {}) or {}
    lines = [_s(data.get("problem_description", ""))]
    for label, key in [
        ("WHAT", "what"),
        ("WHERE", "where"),
        ("WHEN", "when"),
        ("WHO", "who"),
        ("WHY", "why"),
        ("HOW", "how"),
        ("HOW MANY", "how_many"),
    ]:
        val = _s(five_w.get(key, ""))
        if val:
            lines.append(f"{label}: {val}")
    for label, key in [
        ("Standard", "standard_applicable"),
        ("Expected", "expected_situation"),
        ("Observed", "observed_situation"),
    ]:
        val = _s(data.get(key, ""))
        if val:
            lines.append(f"{label}: {val}")

    text = "\n".join(filter(None, lines))
    _w(ws, "B14", text)
    _wrap(ws, "B14")
    _auto_height(ws, 14, text, col_width=27)

    # Photo + file links
    img_bytes, links = _categorise_files(db, d2_step)
    if img_bytes:
        _insert_photo(ws, img_bytes, anchor="G15", width=380, height=180)
    if links:
        _write_links(ws, "G26", links)

    # Is / Is Not
    FACTOR_ROWS = {"product": 21, "time": 22, "lot": 23, "pattern": 24}
    for factor in data.get("is_is_not_factors", []) or []:
        factor_name = _s(factor.get("factor", "")).strip().lower()
        row = FACTOR_ROWS.get(factor_name)
        if not row:
            continue
        is_val = (
            factor.get("is_problem")
            if factor.get("is_problem") is not None
            else factor.get("is_value")
        )
        is_not_val = (
            factor.get("is_not_problem")
            if factor.get("is_not_problem") is not None
            else factor.get("is_not_value")
        )
        _w(ws, f"C{row}", _s(is_val, ""))
        _w(ws, f"E{row}", _s(is_not_val, ""))
        _wrap(ws, f"C{row}")
        _wrap(ws, f"E{row}")


# =============================================================================
# D3 — Containment
# =============================================================================


def _fill_d3(
    ws: Worksheet,
    data: Dict,
    db: Session,
    d3_step: Optional[ReportStep],
) -> None:
    defected: Dict = data.get("defected_part_status", {}) or {}
    _wb(ws, "C27", defected.get("returned", False))
    _wb(ws, "E27", defected.get("isolated", False))
    if defected.get("isolated"):
        _w(ws, "F27", f"Isolated — {_s(defected.get('isolated_location'))}")
        _wrap(ws, "F27")
    _wb(ws, "H27", defected.get("identified", False))
    if defected.get("identified"):
        _w(ws, "I27", f"Identified: {_s(defected.get('identified_method'))}")
        _wrap(ws, "I27")

    LOC_ROWS = {
        "supplier_site": 31,
        "in_transit": 32,
        "production_floor": 33,
        "warehouse": 34,
        "customer_site": 35,
        "others": 36,
    }
    for rd in data.get("suspected_parts_status", []) or []:
        r = LOC_ROWS.get(rd.get("location", ""))
        if r:
            for col, key in [
                ("C", "inventory"),
                ("E", "actions"),
                ("G", "leader"),
                ("K", "results"),
            ]:
                val = _s(rd.get(key))
                _w(ws, f"{col}{r}", val)
                if val:
                    _wrap(ws, f"{col}{r}")

    alert: Dict = data.get("alert_communicated_to", {}) or {}
    _wb(ws, "C37", alert.get("production_shift_leaders", False))
    _wb(ws, "E37", alert.get("warehouse", False))
    _wb(ws, "H37", alert.get("customer_contact", False))
    _wb(ws, "C38", alert.get("quality_control", False))
    _wb(ws, "E38", alert.get("maintenance", False))
    _wb(ws, "H38", alert.get("production_planner", False))

    alert_num = _s(data.get("alert_number", ""))
    _w(
        ws,
        "B38",
        (
            f"Alert # (QRQC log or NCR #): {alert_num}"
            if alert_num
            else "Alert # (QRQC log or NCR #):"
        ),
    )

    restart: Dict = data.get("restart_production", {}) or {}
    for coord, key in [
        ("D40", "when"),
        ("G40", "first_certified_lot"),
        ("K40", "identification"),
        ("D41", "approved_by"),
        ("D42", "method"),
    ]:
        val = _s(restart.get(key))
        _w(ws, coord, val)
        if val:
            _wrap(ws, coord)

    val = _s(data.get("containment_responsible"))
    _w(ws, "C44", val)
    if val:
        _wrap(ws, "C44")

    _, links = _categorise_files(db, d3_step)
    if links:
        _write_links(ws, "K44", links)


# =============================================================================
# D4 — Root cause  (4M + 5 Whys × 2)
# =============================================================================


def _fill_5whys(ws: Worksheet, whys: Dict, start: int) -> None:
    for i in range(5):
        why = (whys or {}).get(f"why_{i+1}", {}) or {}
        base = start + i * 6
        q = _s(why.get("question"))
        a = _s(why.get("answer"))
        _w(ws, f"D{base}", q)
        _w(ws, f"D{base+3}", a)
        if q:
            _wrap(ws, f"D{base}")
            _auto_height(ws, base, q, 50)
        if a:
            _wrap(ws, f"D{base+3}")
            _auto_height(ws, base + 3, a, 50)


def _fill_4m(ws: Worksheet, fourm: Dict, base: int) -> None:
    for i, key in enumerate(["row_1", "row_2", "row_3"]):
        rd = fourm.get(key) or {}
        r = base + 1 + i
        r2 = base + 6 + i
        for col, fk in [("B", "material"), ("D", "method"), ("F", "machine")]:
            val = _s(rd.get(fk))
            _w(ws, f"{col}{r}", val)
            if val:
                _wrap(ws, f"{col}{r}")
        for col, fk in [("B", "manpower"), ("D", "environment")]:
            val = _s(rd.get(fk))
            _w(ws, f"{col}{r2}", val)
            if val:
                _wrap(ws, f"{col}{r2}")
    val = _s(fourm.get("selected_problem"))
    _w(ws, f"J{base+3}", val)
    if val:
        _wrap(ws, f"J{base+3}")


def _fill_d4(
    ws: Worksheet,
    data: Dict,
    db: Session,
    d4_step: Optional[ReportStep],
) -> None:
    _fill_4m(ws, data.get("four_m_occurrence", {}) or {}, 48)
    _fill_5whys(ws, data.get("five_whys_occurrence", {}) or {}, 61)
    rc = data.get("root_cause_occurrence", {}) or {}
    for coord, key in [("B91", "root_cause"), ("E91", "validation_method")]:
        val = _s(rc.get(key))
        _w(ws, coord, val)
        if val:
            _wrap(ws, coord)

    _fill_4m(ws, data.get("four_m_non_detection", {}) or {}, 95)
    _fill_5whys(ws, data.get("five_whys_non_detection", {}) or {}, 106)
    rc2 = data.get("root_cause_non_detection", {}) or {}
    for coord, key in [("B136", "root_cause"), ("E136", "validation_method")]:
        val = _s(rc2.get(key))
        _w(ws, coord, val)
        if val:
            _wrap(ws, coord)

    _, links = _categorise_files(db, d4_step)
    if links:
        _write_links(ws, "B137", links)


# =============================================================================
# D5 + D6 — Corrective actions & implementation
# =============================================================================


def _fill_d5_d6(
    ws: Worksheet,
    d5: Dict,
    d6: Dict,
    db: Session,
    d5_step: Optional[ReportStep],
    d6_step: Optional[ReportStep],
) -> None:
    d6_occ = d6.get("corrective_actions_occurrence", []) or []
    d6_det = d6.get("corrective_actions_detection", []) or []

    for idx, row in enumerate((d5.get("corrective_actions_occurrence", []) or [])[:2]):
        r = 142 + idx
        d6r = d6_occ[idx] if idx < len(d6_occ) else {}
        for col, key in [("B", "action"), ("F", "responsible"), ("G", "due_date")]:
            val = _s(row.get(key))
            _w(ws, f"{col}{r}", val)
            if val:
                _wrap(ws, f"{col}{r}")
        _w(ws, f"I{r}", _s(d6r.get("imp_date", "")))
        val = _s(d6r.get("evidence", ""))
        _w(ws, f"K{r}", val)
        if val:
            _wrap(ws, f"K{r}")

    for idx, row in enumerate((d5.get("corrective_actions_detection", []) or [])[:2]):
        r = 145 + idx
        d6r = d6_det[idx] if idx < len(d6_det) else {}
        for col, key in [("B", "action"), ("F", "responsible"), ("G", "due_date")]:
            val = _s(row.get(key))
            _w(ws, f"{col}{r}", val)
            if val:
                _wrap(ws, f"{col}{r}")
        _w(ws, f"I{r}", _s(d6r.get("imp_date", "")))
        val = _s(d6r.get("evidence", ""))
        _w(ws, f"K{r}", val)
        if val:
            _wrap(ws, f"K{r}")

    mon: Dict = d6.get("monitoring", {}) or {}
    _w(ws, "C149", _s(mon.get("monitoring_interval")))
    _w(ws, "C150", _s(mon.get("pieces_produced")))
    _w(ws, "C151", _s(mon.get("rejection_rate")))

    for idx, item in enumerate((d6.get("checklist", []) or [])[:13]):
        r = 149 + idx
        _wb(ws, f"K{r}", item.get("shift_1", False))
        _wb(ws, f"L{r}", item.get("shift_2", False))
        _wb(ws, f"M{r}", item.get("shift_3", False))

    _w(ws, "F162", _s(mon.get("audited_by")))
    _w(ws, "F163", _s(mon.get("audit_date")))

    _, links5 = _categorise_files(db, d5_step)
    _, links6 = _categorise_files(db, d6_step)
    all_links = links5 + links6
    if all_links:
        _write_links(ws, "K164", all_links)


# =============================================================================
# D7 — Lessons learned / prevention
# =============================================================================


def _fill_d7(
    ws: Worksheet,
    data: Dict,
    db: Session,
    d7_step: Optional[ReportStep],
) -> None:
    for idx, risk in enumerate((data.get("recurrence_risks", []) or [])[:3]):
        r = 168 + idx
        for col, key in [("B", "area_line_product"), ("E", "action_taken")]:
            val = _s(risk.get(key))
            _w(ws, f"{col}{r}", val)
            if val:
                _wrap(ws, f"{col}{r}")
        _wb(ws, f"C{r}", bool(risk.get("similar_risk_present", False)))

    for idx, item in enumerate((data.get("lesson_disseminations", []) or [])[:3]):
        r = 174 + idx
        for col, key in [
            ("B", "audience_team"),
            ("D", "method"),
            ("F", "date"),
            ("G", "owner"),
            ("J", "evidence"),
        ]:
            val = _s(item.get(key))
            _w(ws, f"{col}{r}", val)
            if val:
                _wrap(ws, f"{col}{r}")

    for idx, item in enumerate((data.get("replication_validations", []) or [])[:3]):
        r = 179 + idx
        for col, key in [
            ("B", "line_site"),
            ("C", "action_replicated"),
            ("F", "confirmation_method"),
            ("G", "confirmed_by"),
        ]:
            val = _s(item.get(key))
            _w(ws, f"{col}{r}", val)
            if val:
                _wrap(ws, f"{col}{r}")

    for idx, item in enumerate((data.get("knowledge_base_updates", []) or [])[:3]):
        r = 184 + idx
        for col, key in [
            ("B", "document_type"),
            ("C", "topic_reference"),
            ("F", "owner"),
            ("G", "location_link"),
        ]:
            val = _s(item.get(key))
            _w(ws, f"{col}{r}", val)
            if val:
                _wrap(ws, f"{col}{r}")

    LT_FREQ_COL = {189: "D", 190: "C", 191: "D"}
    for idx, item in enumerate((data.get("long_term_monitoring", []) or [])[:3]):
        r = 189 + idx
        freq_col = LT_FREQ_COL[r]
        for col, key in [
            ("B", "checkpoint_type"),
            (freq_col, "frequency"),
            ("F", "owner"),
            ("G", "start_date"),
            ("L", "notes"),
        ]:
            val = _s(item.get(key))
            _w(ws, f"{col}{r}", val)
            if val:
                _wrap(ws, f"{col}{r}")

    val = _s(data.get("ll_conclusion"))
    _w(ws, "B193", val)
    if val:
        _wrap(ws, "B193")
        _auto_height(ws, 193, val, col_width=80)

    _, links = _categorise_files(db, d7_step)
    if links:
        _write_links(ws, "B194", links)


# =============================================================================
# D8 — Closure
# =============================================================================


def _fill_d8(
    ws: Worksheet,
    data: Dict,
    db: Session,
    d8_step: Optional[ReportStep],
) -> None:
    closure = _s(data.get("closure_statement", ""))
    sigs: Dict = data.get("signatures", {}) or {}
    parts = []
    if sigs.get("closed_by"):
        parts.append(f"Closed by: {_s(sigs['closed_by'])}")
    if sigs.get("closure_date"):
        parts.append(f"Date: {_s(sigs['closure_date'])}")
    if sigs.get("approved_by"):
        parts.append(f"Approved by: {_s(sigs['approved_by'])}")
    sig_line = "  |  ".join(parts)
    full = (
        (closure + "\n\n" + sig_line)
        if (closure and sig_line)
        else (closure or sig_line)
    )
    _w(ws, "B199", full)
    if full:
        _wrap(ws, "B199")
        _auto_height(ws, 199, full, col_width=80)

    _, links = _categorise_files(db, d8_step)
    if links:
        _write_links(ws, "B202", links)


# =============================================================================
# SERVICE
# =============================================================================


class ReportExportService:

    @staticmethod
    def generate_excel(db: Session, report_id: int) -> bytes:
        report = db.query(Report).filter(Report.id == report_id).first()
        if not report:
            raise ValueError(f"Report {report_id} not found")
        complaint: Complaint = report.complaint
        if not complaint:
            raise ValueError("Report has no associated complaint")

        steps: Dict[str, ReportStep] = {s.step_code: s for s in report.steps}

        wb = load_workbook(TEMPLATE_PATH)
        ws = wb["Sheet1"]

        _w(
            ws,
            "B2",
            (
                f"PROBLEM SOLVING - 8D METHOD  |  "
                f"{_s(report.report_number)}  |  "
                f"{_s(complaint.complaint_name)}"
            ),
        )

        _fill_d1(ws, _step_data(steps, "D1"))
        _fill_d2(ws, _step_data(steps, "D2"), db, steps.get("D2"))
        _fill_d3(ws, _step_data(steps, "D3"), db, steps.get("D3"))
        _fill_d4(ws, _step_data(steps, "D4"), db, steps.get("D4"))
        _fill_d5_d6(
            ws,
            _step_data(steps, "D5"),
            _step_data(steps, "D6"),
            db,
            steps.get("D5"),
            steps.get("D6"),
        )
        _fill_d7(ws, _step_data(steps, "D7"), db, steps.get("D7"))
        _fill_d8(ws, _step_data(steps, "D8"), db, steps.get("D8"))
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        return buf.read()

    @staticmethod
    def get_filename(db: Session, report_id: int) -> str:
        report = db.query(Report).filter(Report.id == report_id).first()
        if not report:
            return f"8D_report_{report_id}.xlsx"
        name = _s(getattr(report.complaint, "complaint_name", ""))[:40]
        name = name.replace(" ", "_").replace("/", "-")
        name = _safe_filename(name)
        return f"8D_{_s(report.report_number)}_{name}.xlsx"

    @staticmethod
    # def is_report_ready(report: Optional[Report]) -> bool:
    #         if not report:
    #             return False

    #         step_map: Dict[str, ReportStep] = {s.step_code: s for s in (report.steps or [])}

    #         for code in _STEP_ORDER:
    #             step = step_map.get(code)
    #             if step is None or step.status != "fulfilled":
    #                 return False

    #         return True
    def is_report_ready(report: Optional[Report]) -> bool:
        """
        A report is exportable as long as it exists.
        Steps that are not yet fulfilled will simply export as empty/blank
        sections in the Excel — no hard gate on all-steps-fulfilled.
        """
        return report is not None

    @staticmethod
    def has_export_for_complaint(db: Session, complaint_id: int) -> bool:
        report = db.query(Report).filter(Report.complaint_id == complaint_id).first()
        return ReportExportService.is_report_ready(report)

    @staticmethod
    def get_export_meta_for_complaint(db: Session, complaint_id: int) -> Dict[str, Any]:
        report = db.query(Report).filter(Report.complaint_id == complaint_id).first()

        if not report:
            return {
                "has_export_report": False,
                "export_filename": None,
            }

        ready = ReportExportService.is_report_ready(report)

        return {
            "has_export_report": ready,
            "export_filename": (
                ReportExportService.get_filename(db, report.id) if ready else None
            ),
        }

    @staticmethod
    async def save_to_github(db: Session, report_id: int) -> str:
        """Generate Excel, push to GitHub reports folder, return URL."""
        file_bytes = ReportExportService.generate_excel(db, report_id)
        filename = ReportExportService.get_filename(db, report_id)
        result = await storage.upload_report(file_bytes, filename)
        return result["url"]
